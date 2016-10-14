#coding=utf-8

from openpyxl import load_workbook
import csv
import requests
from Tkinter import Tk
from tkFileDialog import askopenfilename, asksaveasfilename
import re


def load_category_csv():
    result = dict()
    with open('category.csv', 'rb') as csvfile:
        reader = csv.reader(csvfile, delimiter=',')
        for id_erp,id_promo in reader:
            result[id_erp] = id_promo
    return result

category_dict = load_category_csv()

def load_brands():
    response = requests.get('http://www.detmir.ru/api/rest/dictionaries')
    data = response.json()
    result = dict()
    for brand in data['brands']:
        result[brand['title'].replace(' ', '').lower()] = brand['id']
    return result


brands_dict = load_brands()

Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
filename = askopenfilename(filetypes=(('Excel files','.xlsx'),('All types', '*.*'))) # show an "Open" dialog box and return the path to the selected file
print(filename)

wb = load_workbook(filename)

# grab the active worksheet
ws = wb.worksheets[0]

ws2 = wb.create_sheet(title="Result")

regex = re.compile('\(.*?\)')


def copy_values(idx, index, is_sup):
    # 1-уникальный ID ДМ
    val = unicode(ws.cell(row=idx, column=1).value)
    if is_sup:
        val += u'sup'
    ws2.cell(row=index, column=2).value = val

    # 2-Артикул
    ws2.cell(row=index, column=3).value = ws.cell(row=idx, column=2).value

    # 3-Наименование
    if ws.cell(row=idx, column=3).value:
        vals = ws.cell(row=idx, column=3).value.split(u':')
        if is_sup:
            ws2.cell(row=index, column=5).value = vals[0].capitalize()
            ws2.cell(row=index, column=7).value = vals[1]
        elif len(vals) >= 3:
            ws2.cell(row=index, column=5).value = vals[0].capitalize()
            ws2.cell(row=index, column=7).value = vals[2]
        else:
            ws2.cell(row=index, column=5).value = ws.cell(row=idx, column=3).value

    # 4-Категория
    if ws.cell(row=idx, column=4).value:
        id_erp = re.sub(regex, '', ws.cell(row=idx, column=4).value)
        if id_erp in category_dict:
            ws2.cell(row=index, column=4).value = category_dict[id_erp]
        else:
            ws2.cell(row=index, column=4).value = ws.cell(row=idx, column=4).value

    # 7-Бренд
    if ws.cell(row=idx, column=5).value:
        brand = re.sub(regex, '', ws.cell(row=idx, column=5).value)
        if brand and brand.replace(' ', '').lower() in brands_dict:
            ws2.cell(row=index, column=6).value = brands_dict[brand.replace(' ', '').lower()]
        else:
            ws2.cell(row=index, column=6).value = ws.cell(row=idx, column=5).value

    # 10-пол
    if ws.cell(row=idx, column=7).value:
        if ws.cell(row=idx, column=7).value.strip().lower() == u'женский':
            ws2.cell(row=index, column=10).value = u'ж'
        elif ws.cell(row=idx, column=7).value.strip().lower() == u'мужской':
            ws2.cell(row=index, column=10).value = u'м'
        elif ws.cell(row=idx, column=7).value.strip().lower() == u'унисекс':
            ws2.cell(row=index, column=10).value = ' '
        else:
            ws2.cell(row=index, column=10).value = ws.cell(row=idx, column=10).value

    # 13-состав товара
    ws2.cell(row=index, column=16).value = ws.cell(row=idx, column=8).value

    # 8-возрастная группа
    if ws.cell(row=idx, column=6).value:
        if ws.cell(row=idx, column=6).value.strip().lower() == u'от 0 мес.':
            ws2.cell(row=index, column=18).value = u'0'
            ws2.cell(row=index, column=19).value = u'24'
        elif ws.cell(row=idx, column=6).value.strip().lower() == u'от 3 мес.':
            ws2.cell(row=index, column=18).value = u'3'
        elif ws.cell(row=idx, column=6).value.strip().lower() == u'от 6 мес.':
            ws2.cell(row=index, column=18).value = u'6'
        elif ws.cell(row=idx, column=6).value.strip().lower() == u'от 1,5 лет':
            ws2.cell(row=index, column=18).value = u'18'
        elif ws.cell(row=idx, column=6).value.strip().lower() == u'от 2 лет':
            ws2.cell(row=index, column=18).value = u'24'
            ws2.cell(row=index, column=19).value = u'72'
        elif ws.cell(row=idx, column=6).value.strip().lower() == u'от 3 лет':
            ws2.cell(row=index, column=18).value = u'36'
        elif ws.cell(row=idx, column=6).value.strip().lower() == u'от 7 лет':
            ws2.cell(row=index, column=18).value = u'84'
            ws2.cell(row=index, column=19).value = u'144'
        else:
            ws2.cell(row=index, column=18).value = ws.cell(row=idx, column=6).value

    # 15-Коллекция
    ws2.cell(row=index, column=24).value = ws.cell(row=idx, column=10).value + ws.cell(row=idx, column=9).value

    #11 - Страна происхождения
    ws2.cell(row=index, column=17).value = ws.cell(row=idx, column=11).value

    #12 - Ширина
    ws2.cell(row=index, column=12).value = ws.cell(row=idx, column=12).value * 100

    #13 - Высота
    ws2.cell(row=index, column=14).value = ws.cell(row=idx, column=13).value * 100

    #14 - Длина
    ws2.cell(row=index, column=13).value = ws.cell(row=idx, column=14).value * 100

    #15 - Вес
    ws2.cell(row=index, column=15).value = ws.cell(row=idx, column=15).value * 1000

    #Габариты
    ws2.cell(row=index, column=11).value = unicode('{0:.0f}'.format(float(ws.cell(row=idx, column=12).value) * 100)) + u'x'

    # + unicode(round(ws.cell(row=idx, column=13).value * 100.0)) + u'x' + unicode(round(ws.cell(row=idx, column=14).value * 100.0))

# Copy
col_d = ws.columns[0]
i = 1
for idx, c in enumerate(col_d, 1):
    if unicode(ws.cell(row=idx, column=1).value)[-3:] == u'001':
        ws2.cell(row=i, column=1).value = str(i - 1)
        copy_values(idx, i, True)
        i += 1

    ws2.cell(row=i, column=1).value = str(i-1)
    copy_values(idx, i, False)
    i += 1


ws2.cell(row=1, column=1).value = u'н/п'
ws2.cell(row=1, column=2).value = u'уникальный ID ДМ'
ws2.cell(row=1, column=3).value = u'Артикул'
ws2.cell(row=1, column=4).value = u'Категория'
ws2.cell(row=1, column=5).value = u'Наименование'
ws2.cell(row=1, column=6).value = u'Бренд'
ws2.cell(row=1, column=7).value = u'Модель'
ws2.cell(row=1, column=8).value = u'Цвет'
ws2.cell(row=1, column=9).value = u'Описание'
ws2.cell(row=1, column=10).value = u'Пол'
ws2.cell(row=1, column=11).value = u'Габариты'
ws2.cell(row=1, column=12).value = u'Ширина см'
ws2.cell(row=1, column=13).value = u'Длина см'
ws2.cell(row=1, column=14).value = u'Высота см'
ws2.cell(row=1, column=15).value = u'Вес'
ws2.cell(row=1, column=16).value = u'Материал'
ws2.cell(row=1, column=17).value = u'Страна производитель'
ws2.cell(row=1, column=18).value = u'Возрастная группа от'
ws2.cell(row=1, column=19).value = u'Возрастная группа до'
ws2.cell(row=1, column=20).value = u'КГТ'
ws2.cell(row=1, column=21).value = u'Рекомендуем?'
ws2.cell(row=1, column=22).value = u'теги, через запятую'
ws2.cell(row=1, column=23).value = u'Рейтинг'
ws2.cell(row=1, column=24).value = u'Коллекция'
ws2.cell(row=1, column=25).value = u'Цена'
ws2.cell(row=1, column=26).value = u'Ценовой уровень'

# Save the file
wb.save("result.xlsx")
